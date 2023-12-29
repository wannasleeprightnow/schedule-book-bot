import re
from wsgiref.handlers import format_date_time

from openpyxl import load_workbook

wb = load_workbook(filename='temp.xlsx')
ws = wb.active
values = []
for row in ws.iter_rows(max_row=49, max_col=6, values_only=True):
    values.append(list(row))


def format_table(table):
    table = [j for i in table for j in i]
    formated_table = []
    date = " ".join(table[0].split()[-2:]) + "\n"
    for elem in table[1:]:
        if elem is not None:
            elem = "".join(elem) if type(elem) == list else str(elem)
            elem = re.sub(r'[^А-Яа-яё/ ]', "", elem) if len(elem) > 3 else elem
            elem = elem[:-1] if elem[-1] == "/" else elem
            elem = elem.strip()
        formated_table.append(elem)
    #         formated_table.append("---")
    # return date + '\n'.join(["\t\t\t".join(formated_table[i: i + 6]) for i in range(5, len(formated_table[5:]), 6)])
    return [formated_table[i: i + 6] for i in range(5, len(formated_table[5:]), 6)]


def diff_classes(table):
    limits = [table.index(i) for i in table if i[0] is None and not i[1] is None]
    limits.append(len(table))
    parallels = {}
    for limit in range(1, len(limits)):
        parallel = table[limits[limit - 1]: limits[limit]]
        for i in range(1, len(parallel[0])):
            if not parallel[0][i] is None:
                parallels[parallel[0][i]] = []
                for j in range(1, len(parallel)):
                    if not parallel[j][i] is None:
                        parallels[parallel[0][i]].append(parallel[j][i])
    return parallels


print(diff_classes(format_table(values)))
