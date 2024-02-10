from dataclasses import dataclass
import datetime
from itertools import chain
import re

from openpyxl import load_workbook

from config import FormatsConfig, MailConfig


@dataclass
class ParsedSchedule:
    number_letter_grade: str
    lessons_date: datetime.date
    lessons: list[str]


class ScheduleExcelParser:
    def __init__(self):
        self.active_workbook = load_workbook(MailConfig.SAVE_FILENAME).active
        self._table = self._get_rows()
        self.date = self._get_date()
        self._formated_table = self._format_table()

    def _get_rows(self):
        rows = []
        for row in self.active_workbook.iter_rows(max_col=6, values_only=True):
            if row == ("---"):
                break
            rows.append(list(row))
        return rows

    def _get_date(self):
        date_ = re.sub(r"[^0-9.]", "", self._table[0][0].strip())
        date_ = date_.replace(".23", ".2023").replace(".24", ".2024")
        return datetime.datetime.strptime(
            date_, FormatsConfig.DATE_FORMAT).date()

    def _format_elem(self, elem):
        if elem is not None:
            elem = str(elem).strip()
            elem = re.sub(r"[^А-Яа-яё/ ]", "", elem) if len(elem) > 3 else elem
            elem = elem.strip()
            elem = elem[:-2] if elem[-1] == "/" else elem
        return elem

    def _format_table(self):
        table = list(chain.from_iterable(self._table))
        formated_table = [self._format_elem(elem) for elem in table[1:]]
        table = [
            formated_table[i : i + 6]
            for i in range(5, len(formated_table[5:]), 6)
        ]
        return table

    def _get_indencies(self):
        indencies = []
        for i in self._formated_table:
            if i[0] is None and not i[1] is None and i[1][-1] == "а":
                indencies.append(self._formated_table.index(i))
        return indencies + [len(self._formated_table)]

    def _get_grade_schedule(self, grade, parallel):
        return ParsedSchedule(
            number_letter_grade=parallel[0][grade],
            lessons_date=self.date,
            lessons=[
                parallel[j][grade]
                for j in range(1, len(parallel))
                if not parallel[j][grade] is None
            ],
        )

    def correlate_grades_schedule(self) -> list[ParsedSchedule]:
        schedules = []
        parallels_indicies = self._get_indencies()
        for i in range(1, len(parallels_indicies)):
            parallel = self._formated_table[
                parallels_indicies[i - 1] : parallels_indicies[i]
            ]
            for grade in range(1, len(parallel[0])):
                if not parallel[0][grade] is None:
                    schedule = self._get_grade_schedule(grade, parallel)
                    if schedule.lessons:
                        schedules.append(schedule)
        return schedules
